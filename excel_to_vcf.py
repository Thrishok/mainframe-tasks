"""
Convert results.xlsx -> contacts.vcf (a single vCard file with all contacts).

Usage:
    python excel_to_vcf.py
    python excel_to_vcf.py --file results.xlsx --country 91 --prefix "Lead - "

Then transfer contacts.vcf to your phone (email it to yourself, Google Drive,
or USB), open it, and tap "Import" / "Add all contacts". After that, open
WhatsApp -> New broadcast -> select these contacts.
"""

import argparse
import re
from pathlib import Path

from openpyxl import load_workbook


def normalize_phone(raw: str, default_country: str) -> str | None:
    if not raw:
        return None
    s = str(raw).strip()
    has_plus = s.startswith("+")
    digits = re.sub(r"\D", "", s)
    if not digits:
        return None
    if has_plus:
        return "+" + digits
    digits = digits.lstrip("0")
    if not digits.startswith(default_country):
        digits = default_country + digits
    return "+" + digits


def make_vcard(name: str, phone: str) -> str:
    # vCard 3.0 — widely supported by Android, iOS, Google Contacts.
    # Escape semicolons and commas in name.
    safe_name = name.replace(";", ",").replace("\n", " ").strip()
    return (
        "BEGIN:VCARD\r\n"
        "VERSION:3.0\r\n"
        f"FN:{safe_name}\r\n"
        f"N:{safe_name};;;;\r\n"
        f"TEL;TYPE=CELL:{phone}\r\n"
        "END:VCARD\r\n"
    )


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", default="results.xlsx", help="Input Excel file")
    ap.add_argument("--country", default="91", help="Default country code (default 91 India)")
    ap.add_argument("--out", default="contacts.vcf", help="Output vCard file")
    ap.add_argument("--prefix", default="Lead - ", help="Prefix added to each contact name (helps you find them)")
    args = ap.parse_args()

    src = Path(args.file)
    if not src.exists():
        print(f"File not found: {src}")
        return

    wb = load_workbook(src, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    name_idx = headers.index("name") if "name" in headers else 0
    phone_idx = headers.index("phone") if "phone" in headers else 1

    seen = set()
    cards = []
    skipped_no_phone = 0
    skipped_dup = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = (row[name_idx] or "").strip() if row[name_idx] else ""
        phone = normalize_phone(row[phone_idx], args.country)
        if not phone:
            skipped_no_phone += 1
            continue
        if phone in seen:
            skipped_dup += 1
            continue
        seen.add(phone)
        full_name = f"{args.prefix}{name}" if name else f"{args.prefix}{phone}"
        cards.append(make_vcard(full_name, phone))

    out = Path(args.out)
    out.write_text("".join(cards), encoding="utf-8")

    print(f"Wrote {len(cards)} contacts to {out}")
    print(f"  skipped (no phone)   : {skipped_no_phone}")
    print(f"  skipped (duplicates) : {skipped_dup}")
    print()
    print("Next steps:")
    print(f"  1. Email/Drive/USB-transfer {out} to your phone.")
    print("  2. Open it on the phone -> Import all contacts.")
    print("  3. WhatsApp -> menu -> New broadcast -> select these contacts.")


if __name__ == "__main__":
    main()
